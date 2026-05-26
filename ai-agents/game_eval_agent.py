import os
import sys
import glob
import re
import shutil
import urllib.parse
import urllib.request
from urllib.error import HTTPError, URLError
from datetime import datetime
import openpyxl
import xlrd

# Fix Windows console encoding issues for Unicode paths
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")
if hasattr(sys.stderr, "reconfigure"):
    sys.stderr.reconfigure(encoding="utf-8")

class GameEvalAgent:
    def __init__(self, project_name="Unknown Project", genre="Casual", competitors=None):
        self.project_name = project_name
        self.genre = genre
        self.competitors = competitors or []
        self.raw_content = ""
        self.extracted_framework = {} # Format: {Category: {"comment": str, "score": float}}
        self.web_research_results = []

    def is_zip_file(self, file_path):
        """Checks if a file has a ZIP signature (e.g. .xlsx renamed to .xls)."""
        try:
            with open(file_path, "rb") as f:
                sig = f.read(4)
                return sig == b"PK\x03\x04"
        except Exception:
            return False

    def is_valid_row(self, heading, desc):
        """Filters out Excel metadata, formulas, score guidelines, and instruction rows."""
        if not heading or not desc:
            return False
        # Skip formulas
        if str(heading).startswith("=") or str(desc).startswith("="):
            return False
        # Skip pure numeric headers (score legend indicators like 15.0, 20.0, etc.)
        if re.match(r'^\d+(\.\d+)?$', str(heading)) or str(heading).replace('.', '').isdigit():
            return False
        # Skip score guidelines
        desc_lower = str(desc).lower()
        if "point" in desc_lower or "1~10" in desc_lower or "1-10" in desc_lower:
            return False
        # Skip common metadata fields that are parsed separately
        heading_lower = str(heading).lower()
        metadata_keywords = [
            "game title", "tên game", "thể loại", "genre", "đối thủ", "competitors", 
            "người test", "ngày test", "tên tester", "tuổi tester", "ngày bắt đầu", 
            "link video", "chủ đề", "link", "giới tính tester", "ngày test", "máy chủ test",
            "độ tuổi", "dung lượng", "icon", "đã publish"
        ]
        if any(x in heading_lower for x in metadata_keywords):
            return False
        return True

    def load_from_xlsx(self, file_path):
        """Unified method to read evaluation content and scores from both .xlsx and legacy .xls files."""
        if not os.path.exists(file_path):
            print(f"File not found: {file_path}")
            return False
            
        is_xlsx = True
        # If it ends with .xls, check if it's actually xlsx renamed (ZIP signature)
        if file_path.lower().endswith(".xls"):
            if self.is_zip_file(file_path):
                is_xlsx = True
            else:
                is_xlsx = False
                
        try:
            if is_xlsx:
                # Open with openpyxl
                temp_path = None
                if file_path.lower().endswith(".xls"):
                    # Create a temp .xlsx copy so openpyxl doesn't complain about extension
                    temp_path = file_path + ".temp.xlsx"
                    shutil.copyfile(file_path, temp_path)
                    wb = openpyxl.load_workbook(temp_path, read_only=True)
                else:
                    wb = openpyxl.load_workbook(file_path, read_only=True)
                
                sheet_names = wb.sheetnames
                print(f"Loaded Excel sheets: {sheet_names} from {os.path.basename(file_path)} (using openpyxl)")
                
                # Helper to extract cell values from a sheet as nested list
                def get_rows(sname):
                    ws = wb[sname]
                    return [[cell for cell in row] for row in ws.iter_rows(values_only=True)]
                    
                # Parsing logic based on sheet names
                if "Testing Report" in sheet_names and "SWOT" in sheet_names:
                    self.parse_format_report_swot(get_rows("Testing Report"), get_rows("SWOT"), file_path)
                elif "Games Info" in sheet_names and "Sheet1" in sheet_names:
                    self.parse_format_games_info(get_rows("Games Info"), get_rows("Sheet1"), file_path)
                elif "Tổng quan" in sheet_names and "Chi tiết" in sheet_names:
                    detail_rows = get_rows("Chi tiết") if "Chi tiết" in sheet_names else None
                    self.parse_format_overview_detail(get_rows("Tổng quan"), detail_rows, file_path)
                else:
                    ws = wb.active
                    active_rows = [[cell for cell in row] for row in ws.iter_rows(values_only=True)]
                    self.parse_format_generic(active_rows, file_path)
                    
                wb.close()
                if temp_path and os.path.exists(temp_path):
                    os.remove(temp_path)
            else:
                # Open legacy .xls with xlrd
                wb = xlrd.open_workbook(file_path)
                sheet_names = wb.sheet_names()
                print(f"Loaded Excel sheets: {sheet_names} from {os.path.basename(file_path)} (using xlrd)")
                
                def get_rows(sname):
                    sheet = wb.sheet_by_name(sname)
                    rows = []
                    for r in range(sheet.nrows):
                        rows.append(sheet.row_values(r))
                    return rows
                    
                if "Testing Report" in sheet_names and "SWOT" in sheet_names:
                    self.parse_format_report_swot(get_rows("Testing Report"), get_rows("SWOT"), file_path)
                elif "Games Info" in sheet_names and "Sheet1" in sheet_names:
                    self.parse_format_games_info(get_rows("Games Info"), get_rows("Sheet1"), file_path)
                elif "Tổng quan" in sheet_names and "Chi tiết" in sheet_names:
                    detail_rows = get_rows("Chi tiết") if "Chi tiết" in sheet_names else None
                    self.parse_format_overview_detail(get_rows("Tổng quan"), detail_rows, file_path)
                else:
                    sheet = wb.sheet_by_index(0)
                    active_rows = []
                    for r in range(sheet.nrows):
                        active_rows.append(sheet.row_values(r))
                    self.parse_format_generic(active_rows, file_path)
            return True
        except Exception as e:
            print(f"Error reading Excel file {os.path.basename(file_path)}: {e}")
            import traceback
            traceback.print_exc()
            return False

    def parse_format_report_swot(self, report_rows, swot_rows, file_path):
        """Format 1: Testing Report + SWOT sheets (e.g. Pokemon Rework, HLV Đỉnh Cao, Càn Quét Tam Quốc)."""
        print("Using parser template: Testing Report & SWOT")
        
        project_name = ""
        genre = ""
        competitors = []
        framework = {}
        
        # 1. Read Testing Report rows
        for r in report_rows:
            clean_row = [str(cell).strip() if cell is not None else "" for cell in r]
            if len(clean_row) < 2:
                continue
            
            col_1 = clean_row[1].lower()
            # Extract game metadata
            if "game title" in col_1 or "tên game" in col_1:
                if len(clean_row) > 3 and clean_row[3]:
                    project_name = clean_row[3]
            elif "thể loại" in col_1 or "genre" in col_1:
                if len(clean_row) > 3 and clean_row[3]:
                    genre = clean_row[3]
            elif "đối thủ" in col_1 or "game tương tự" in col_1:
                if len(clean_row) > 3 and clean_row[3]:
                    competitors = [c.strip() for c in clean_row[3].split(",") if c.strip()]
            # Extract evaluations
            elif clean_row[1] and len(clean_row) > 3 and clean_row[3]:
                heading = clean_row[1]
                desc = clean_row[3]
                # Filter out garbage
                if self.is_valid_row(heading, desc):
                    # Skip helper texts
                    if len(desc) > 10 and not desc.startswith("Yêu cầu") and not desc.startswith("Ví dụ"):
                        framework[heading] = {"comment": desc, "score": 0.0}

        # 2. Read SWOT rows
        current_section = ""
        for r in swot_rows:
            clean_row = [str(cell).strip() if cell is not None else "" for cell in r]
            if not clean_row:
                continue
            
            col_0 = clean_row[0].upper()
            if "STRENGTH" in col_0:
                current_section = "Strengths (Điểm mạnh)"
            elif "WEAKNESS" in col_0:
                current_section = "Weaknesses (Điểm yếu)"
            elif "OPPORTUNIT" in col_0:
                current_section = "Opportunities (Cơ hội)"
            elif "THREAT" in col_0:
                current_section = "Threats (Thách thức)"
                
            if current_section and len(clean_row) > 1 and clean_row[1]:
                content = clean_row[1]
                # Skip instructional row questions
                if len(content) > 15 and "What are" not in content and "List down" not in content and "Why is this" not in content:
                    framework[current_section] = {"comment": content, "score": 0.0}

        self.project_name = project_name
        self.genre = genre
        self.competitors = competitors
        self.extracted_framework = framework
        self.raw_content = " ".join([f"{k}: {v['comment']}" for k, v in framework.items()])

    def parse_format_games_info(self, info_rows, sheet1_rows, file_path):
        """Format 2: Games Info + Sheet1 sheets (e.g. Solo Leveling, Battle of Balls)."""
        print("Using parser template: Games Info & Sheet1")
        
        project_name = ""
        genre = ""
        competitors = []
        framework = {}
        
        # 1. Read Games Info rows
        swot_comment = ""
        eval_comment = ""
        
        for r in info_rows:
            clean_row = [str(cell).strip() if cell is not None else "" for cell in r]
            if len(clean_row) < 2:
                continue
            
            col_0 = clean_row[0].lower()
            col_1 = clean_row[1]
            
            if "tên game" in col_0 or "game title" in col_0:
                project_name = col_1
            elif "genre" in col_0 or "thể loại" in col_0:
                genre = col_1
            elif "đối thủ" in col_0 or "competitors" in col_0:
                competitors = [c.strip() for c in col_1.split(",") if c.strip()]
            elif "swot" in col_0:
                swot_comment = col_1
            elif "đánh giá" in col_0 or "evaluation" in col_0:
                eval_comment = col_1

        # Process SWOT
        if swot_comment:
            swot_sections = re.split(r'(\d+/\s*(?:Strengh|Weakness|Oppotunities|Threats):?)', swot_comment)
            if len(swot_sections) > 1:
                for i in range(1, len(swot_sections), 2):
                    title = swot_sections[i].strip()
                    body = swot_sections[i+1].strip() if i+1 < len(swot_sections) else ""
                    if body:
                        title_clean = "SWOT - " + re.sub(r'[\d/\s:]', '', title).capitalize()
                        framework[title_clean] = {"comment": body, "score": 0.0}
            else:
                framework["SWOT Analysis"] = {"comment": swot_comment, "score": 0.0}
                
        if eval_comment:
            score = 0.0
            match = re.search(r'(\d+(?:\.\d+)?)\s*/\s*10', eval_comment)
            if match:
                score = float(match.group(1)) / 2.0 # convert scale 10 to scale 5
            framework["Tổng kết đánh giá"] = {"comment": eval_comment, "score": score}

        # 2. Read Sheet1 (detailed features)
        for r in sheet1_rows:
            clean_row = [str(cell).strip() if cell is not None else "" for cell in r]
            if len(clean_row) > 3 and clean_row[1] and clean_row[3]:
                feat = clean_row[1]
                desc = clean_row[3]
                if self.is_valid_row(feat, desc):
                    if len(desc) > 15 and feat != "Kỹ năng" and feat != "Thời trang" and "Mô tả" not in desc:
                        framework[feat] = {"comment": desc, "score": 0.0}

        self.project_name = project_name
        self.genre = genre
        self.competitors = competitors
        self.extracted_framework = framework
        self.raw_content = " ".join([f"{k}: {v['comment']}" for k, v in framework.items()])

    def parse_format_overview_detail(self, overview_rows, detail_rows, file_path):
        """Format 3: Tổng quan + Chi tiết sheets (e.g. Ultimate Football Club)."""
        print("Using parser template: Tổng quan & Chi tiết")
        
        project_name = ""
        genre = "Giả lập bóng đá - Quản lý"
        competitors = []
        framework = {}
        
        # Fallback game name from filename
        base_name = os.path.basename(file_path)
        project_name = re.sub(r'(_notes|_eval|_review|_evaluation|sample|đánh giá game|\[.*?\]|\(.*?\))', '', base_name, flags=re.IGNORECASE)
        project_name = project_name.replace(".xlsx", "").replace(".xls", "").replace("_", " ").replace("-", " ").strip()
        project_name = project_name.title() if project_name else "Ultimate Football Club"

        # 1. Read overview sheet rows
        for r in overview_rows:
            clean_row = [str(cell).strip() if cell is not None else "" for cell in r]
            if len(clean_row) > 6 and clean_row[3] and clean_row[6]:
                item = clean_row[3]
                comment = clean_row[6]
                if self.is_valid_row(item, comment):
                    score = 0.0
                    if len(clean_row) > 7 and clean_row[7]:
                        try:
                            score = float(clean_row[7])
                        except ValueError:
                            score = 0.0
                    framework[item] = {"comment": comment, "score": score}

        # 2. Read details sheet rows for strengths & weaknesses
        if detail_rows:
            for r in detail_rows:
                clean_row = [str(cell).strip() if cell is not None else "" for cell in r]
                for cell in clean_row:
                    if cell.startswith("Strengths:") or cell.startswith("Weaknesses:"):
                        lines = cell.split("\n")
                        header = lines[0].strip()
                        body = "\n".join(lines[1:]).strip()
                        if body:
                            framework[header] = {"comment": body, "score": 0.0}

        self.project_name = project_name
        self.genre = genre
        self.competitors = competitors
        self.extracted_framework = framework
        self.raw_content = " ".join([f"{k}: {v['comment']}" for k, v in framework.items()])

    def parse_format_generic(self, rows, file_path):
        """Format 4: Generic key-comment-score table (Standard fallback)."""
        print("Using parser template: Generic Fallback Table")
        
        project_name = ""
        genre = ""
        competitors = []
        framework = {}
        
        clean_rows = []
        for r in rows:
            if any(cell is not None for cell in r):
                clean_rows.append([str(cell).strip() if cell is not None else "" for cell in r])
                
        if not clean_rows:
            return
            
        headers = [h.lower() for h in clean_rows[0]]
        
        category_idx = 0
        comment_idx = 1 if len(clean_rows[0]) > 1 else 0
        score_idx = 2 if len(clean_rows[0]) > 2 else -1
        
        # Check headers
        for idx, h in enumerate(headers):
            if any(k in h for k in ["hạng mục", "tiêu chí", "category", "criteria", "tên", "name"]):
                category_idx = idx
            elif any(k in h for k in ["nhận xét", "chi tiết", "comment", "review", "evaluation", "description", "note"]):
                comment_idx = idx
            elif any(k in h for k in ["điểm", "score", "rating"]):
                score_idx = idx
                
        # Skip header row if matches
        keywords = ["hạng mục", "tiêu chí", "nhận xét", "chi tiết", "điểm", "category", "comment", "score", "criteria", "rating"]
        is_header = any(any(kw in h for kw in keywords) for h in headers)
        start_row = 1 if is_header else 0
        
        for r in clean_rows[start_row:]:
            if len(r) <= category_idx or not r[category_idx]:
                continue
            category = r[category_idx]
            comment = r[comment_idx] if len(r) > comment_idx else ""
            score = 0.0
            if score_idx != -1 and len(r) > score_idx and r[score_idx]:
                try:
                    score = float(r[score_idx])
                except ValueError:
                    score = 0.0
                    
            if any(k in category.lower() for k in ["dự án", "tên game", "project name", "game title"]):
                project_name = comment or category.split(":")[-1].strip()
                continue
            if any(k in category.lower() for k in ["thể loại", "genre"]):
                genre = comment or category.split(":")[-1].strip()
                continue
            if any(k in category.lower() for k in ["đối thủ", "competitors"]):
                competitors = [c.strip() for c in (comment or category.split(":")[-1]).split(",") if c.strip()]
                continue
                
            if self.is_valid_row(category, comment):
                framework[category] = {"comment": comment, "score": score}
            
        # Fallback project name from file name
        if not project_name:
            base_name = os.path.basename(file_path)
            name_without_ext = os.path.splitext(base_name)[0]
            clean_name = re.sub(r'(_notes|_eval|_review|_evaluation|sample|đánh giá)', '', name_without_ext, flags=re.IGNORECASE)
            clean_name = clean_name.replace("_", " ").replace("-", " ").strip()
            project_name = clean_name.title() if clean_name else name_without_ext

        self.project_name = project_name
        self.genre = genre or "Casual"
        self.competitors = competitors
        self.extracted_framework = framework
        self.raw_content = " ".join([f"{k}: {v['comment']}" for k, v in framework.items()])

    def extract_categories_and_criteria(self):
        """Returns the extracted framework (categories and criteria descriptions)."""
        if not self.extracted_framework:
            return {
                "Core Loop": "Gameplay flow and mechanics",
                "Monetization": "Revenue strategy",
                "Visual/UX": "Graphic and Interface quality",
                "Retention": "User engagement factors",
                "USP": "Unique Selling Point"
            }
        return {k: v["comment"] for k, v in self.extracted_framework.items()}

    def _fetch_search_page(self, query, timeout=15):
        """Fetches a search page HTML for a given query using Bing web search."""
        user_agent = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0 Safari/537.36"
        url = f"https://www.bing.com/search?q={urllib.parse.quote_plus(query)}"
        req = urllib.request.Request(url, headers={"User-Agent": user_agent})
        try:
            with urllib.request.urlopen(req, timeout=timeout) as response:
                html = response.read().decode("utf-8", errors="ignore")
                return html
        except (HTTPError, URLError, TimeoutError) as exc:
            print(f"Warning: internet search failed for '{query}': {exc}")
            return ""
        except Exception as exc:
            print(f"Warning: unexpected error during internet search: {exc}")
            return ""

    def _parse_bing_search_results(self, html, max_results):
        """Extracts titles and snippets from Bing search result HTML."""
        results = []
        if not html:
            return results

        pattern = re.compile(r'<li class="b_algo".*?<h2><a href="([^"]+)".*?>(.*?)</a>.*?<p>(.*?)</p>', re.S)
        for match in pattern.finditer(html):
            title = re.sub(r'<.*?>', '', match.group(2)).strip()
            snippet = re.sub(r'<.*?>', '', match.group(3)).strip()
            if title and snippet:
                results.append(f"{title} — {snippet}")
            if len(results) >= max_results:
                break

        if not results:
            fallback_pattern = re.compile(r'<h2><a href="([^"]+)".*?>(.*?)</a>.*?<p>(.*?)</p>', re.S)
            for match in fallback_pattern.finditer(html):
                title = re.sub(r'<.*?>', '', match.group(2)).strip()
                snippet = re.sub(r'<.*?>', '', match.group(3)).strip()
                if title and snippet:
                    results.append(f"{title} — {snippet}")
                if len(results) >= max_results:
                    break

        return results[:max_results]

    def search_web_info(self, max_results=4):
        """Searches the web for game-specific information and returns summarized search results."""
        query_terms = [self.project_name, self.genre, "game review", "gameplay", "analysis"]
        if self.competitors:
            query_terms.append("competitors")
            query_terms.extend(self.competitors[:2])
        query = " ".join([term for term in query_terms if term])

        print(f"AI Agent searching the web for: {query}")
        html = self._fetch_search_page(query)
        results = self._parse_bing_search_results(html, max_results)
        self.web_research_results = results
        if not results:
            print("No web search results could be extracted. Falling back to internal research only.")
        return results

    def research(self):
        """Simulates market research for the project and augments it with online search findings."""
        print(f"AI Agent starting research for: {self.project_name} ({self.genre})")
        research_steps = [
            f"Tìm kiếm dữ liệu đánh giá và lối chơi của {self.project_name}...",
            f"Phân tích xu hướng thị trường game thể loại '{self.genre}' năm 2026...",
        ]
        if self.competitors:
            research_steps.append(f"So sánh các tính năng cốt lõi với đối thủ: {', '.join(self.competitors)}...")
        else:
            research_steps.append("Không phát hiện đối thủ cạnh tranh cụ thể để so sánh.")

        web_results = self.search_web_info(max_results=4)
        if web_results:
            research_steps.append("Tổng hợp thông tin mở rộng từ Internet:")
            for result in web_results:
                research_steps.append(result)
        else:
            research_steps.append("Không thể thu thập thông tin Internet trong bước này.")

        return research_steps

    def generate_scorecard(self, framework=None):
        """Generates scores based on the extracted Excel scores or keyword sentiment analysis."""
        target_framework = framework or self.extracted_framework
        scorecard = {}
        
        for k, v in target_framework.items():
            # If Excel file has a valid score > 0, use it
            if isinstance(v, dict) and v.get("score", 0.0) > 0.0:
                scorecard[k] = v["score"]
            else:
                # Simulated score based on text analysis
                comment_text = v["comment"].lower() if isinstance(v, dict) else str(v).lower()
                base_score = 3.5
                if any(w in comment_text for w in ["tốt", "cuốn", "mượt", "đẹp", "xuất sắc", "great", "excellent", "good", "strengths"]):
                    base_score += 0.8
                if any(w in comment_text for w in ["lạm dụng", "tệ", "lỗi", "giật", "chậm", "bad", "issue", "poor", "weakness"]):
                    base_score -= 0.8
                # Normalize values between 1.0 and 5.0
                scorecard[k] = round(min(5.0, max(1.0, base_score)), 1)
                
        return scorecard

    def write_slides_report(self, scorecard, research_data, output_dir):
        """Writes a beautifully formatted slides-ready Markdown report."""
        os.makedirs(output_dir, exist_ok=True)
        # Clean file name
        safe_project_name = re.sub(r'[\\/*?:"<>|]', "", self.project_name).replace(" ", "_")
        report_path = os.path.join(output_dir, f"Evaluation_{safe_project_name}_{datetime.now().strftime('%Y%m%d')}.md")
        
        avg_score = round(sum(scorecard.values()) / len(scorecard), 2) if scorecard else 0.0
        
        with open(report_path, "w", encoding="utf-8") as f:
            # Main Title
            f.write(f"# Báo cáo Đánh giá sản phẩm: {self.project_name}\n")
            f.write(f"**Ngày tổng hợp:** {datetime.now().strftime('%Y-%m-%d')} | **Thể loại:** {self.genre}\n\n")
            
            # Slide 1: Overview
            f.write("---\n")
            f.write("## Slide 1: Tổng quan dự án\n")
            f.write(f"- **Tên sản phẩm:** {self.project_name}\n")
            f.write(f"- **Thể loại:** {self.genre}\n")
            f.write(f"- **Đối thủ chính:** {', '.join(self.competitors) if self.competitors else 'Không có (N/A)'}\n")
            f.write(f"- **Chỉ số đánh giá trung bình:** {avg_score}/5.0\n")
            f.write("- **Tóm tắt nhanh:** Báo cáo được phân tích tự động từ dữ liệu ghi chú Excel và phân tích cạnh tranh của AI Agent.\n\n")
            
            # Slide 2: Scorecard
            f.write("---\n")
            f.write("## Slide 2: Bảng điểm chất lượng (Scorecard)\n")
            f.write(f"### Điểm trung bình: {avg_score} / 5.0\n\n")
            for criteria, score in scorecard.items():
                short_criteria = criteria[:30] + "..." if len(criteria) > 30 else criteria
                bar = "█" * int(score * 2) + "░" * (10 - int(score * 2))
                f.write(f"- **{short_criteria}**: {score}/5.0 | {bar}\n")
            f.write("\n")
            
            # Slide 3: Detailed comments part 1
            f.write("---\n")
            f.write("## Slide 3: Phân tích Chi tiết Hạng mục (Phần 1)\n")
            items = list(self.extracted_framework.items())
            mid = (len(items) + 1) // 2
            for k, v in items[:mid]:
                score = scorecard.get(k, 0.0)
                f.write(f"### 📍 {k} ({score}/5.0)\n")
                f.write(f"{v['comment']}\n\n")
                
            # Slide 4: Detailed comments part 2
            if len(items) > mid:
                f.write("---\n")
                f.write("## Slide 4: Phân tích Chi tiết Hạng mục (Phần 2)\n")
                for k, v in items[mid:]:
                    score = scorecard.get(k, 0.0)
                    f.write(f"### 📍 {k} ({score}/5.0)\n")
                    f.write(f"{v['comment']}\n\n")
            
            # Slide 5: Research & SWOT
            f.write("---\n")
            f.write("## Slide 5: Kết quả Nghiên cứu Đối thủ & Xu hướng\n")
            for step in research_data:
                f.write(f"- {step}\n")
            if self.web_research_results:
                f.write("\n### Thông tin tham khảo Internet:\n")
                for item in self.web_research_results:
                    f.write(f"- {item}\n")
            f.write("\n### Nhận định chiến lược từ AI:\n")
            f.write("- Cần tập trung cải thiện các điểm hạn chế đã ghi nhận trong phần phân tích chi tiết.\n")
            f.write("- Chuẩn bị các phương án vận hành linh hoạt dựa theo đặc thù thị trường Việt Nam.\n\n")
            
            # Slide 6: Action Items
            f.write("---\n")
            f.write("## Slide 6: Đề xuất Hành động tiếp theo (Action Items)\n")
            f.write("1. **Tối ưu hóa sản phẩm:** Sửa đổi các lỗi thiết kế UI/UX và cơ chế game chưa cân bằng.\n")
            f.write("2. **Định hình mô hình kinh doanh:** Điều chỉnh chiến lược nạp IAP và quảng cáo để không ảnh hưởng xấu tới trải nghiệm người dùng.\n")
            f.write("3. **Xuất PowerPoint:** Sử dụng cấu trúc slide Markdown này để đóng gói thành file PPTX thuyết trình chính thức.\n")
            
        print(f"Report successfully generated at: {report_path}")
        return report_path

def main():
    import argparse
    parser = argparse.ArgumentParser(description="AI Game Evaluation Agent")
    parser.add_argument("--game", type=str, help="Tên game cần đánh giá")
    parser.add_argument("--genre", type=str, default="Casual", help="Thể loại game")
    parser.add_argument("--info", type=str, default="", help="Thông tin chi tiết / ghi chú của game")
    parser.add_argument("--competitors", type=str, default="", help="Danh sách đối thủ cạnh tranh (phân cách bằng dấu phẩy)")
    parser.add_argument("--criteria", type=str, default="", help="Danh sách các tiêu chí đánh giá (phân cách bằng dấu phẩy)")
    args = parser.parse_args()

    # Setup paths
    workspace_dir = "c:/Users/Admin/Desktop/agent AI"
    doc_dir = os.path.join(workspace_dir, "workflow-nhan/doc")
    output_dir = os.path.join(workspace_dir, "workflow-nhan/Đánh giá của AI")
    
    os.makedirs(doc_dir, exist_ok=True)
    os.makedirs(output_dir, exist_ok=True)
    
    # Custom game CLI Mode
    if args.game:
        print(f"Running custom evaluation for Game: {args.game}")
        competitors = [c.strip() for c in args.competitors.split(",") if c.strip()]
        criteria = [c.strip() for c in args.criteria.split(",") if c.strip()]
        if not criteria:
            criteria = ["Core Loop", "Monetization", "Visual/UX", "Retention", "USP"]
            
        agent = GameEvalAgent(project_name=args.game, genre=args.genre, competitors=competitors)
        
        info_lower = args.info.lower()
        framework = {}
        for cr in criteria:
            cr_lower = cr.lower()
            comment = ""
            
            # Simple heuristics for comment generation based on info
            if "core" in cr_lower or "loop" in cr_lower:
                if any(x in info_lower for x in ["gameplay", "chơi", "cơ chế", "loop"]):
                    comment = f"Cơ chế gameplay cốt lõi tập trung vào trải nghiệm người dùng như mô tả: '{args.info}'. Trải nghiệm vòng lặp chính (Core Loop) mượt mà và dễ làm quen."
                else:
                    comment = f"Vòng lặp chính của game {args.game} được thiết kế hợp lý, nhịp độ vừa phải, giúp người chơi dễ tiếp cận nhưng vẫn giữ được độ sâu thử thách ở các giai đoạn sau."
            elif "monet" in cr_lower:
                if any(x in info_lower for x in ["nạp", "tiền", "mua", "ads", "quảng cáo", "iap"]):
                    comment = f"Mô hình kinh doanh kết hợp: '{args.info}'. Cần chú ý cân bằng giữa doanh thu từ IAP/quảng cáo với trải nghiệm người dùng để tránh cảm giác Pay-to-Win."
                else:
                    comment = f"Áp dụng mô hình Hybrid Monetization (kết hợp IAP và quảng cáo xen kẽ). Các điểm chạm mua hàng được sắp đặt tự nhiên, giảm thiểu sự khó chịu cho người chơi miễn phí."
            elif "visual" in cr_lower or "ux" in cr_lower or "đồ họa" in cr_lower:
                if any(x in info_lower for x in ["đồ họa", "hình ảnh", "art", "giao diện", "ui", "ux"]):
                    comment = f"Thiết kế mỹ thuật và giao diện chất lượng cao: '{args.info}'. Visual và hiệu ứng chuyển cảnh tạo ấn tượng thị giác tốt cho người chơi."
                else:
                    comment = f"Phong cách nghệ thuật hiện đại, màu sắc hài hòa và giao diện người dùng tối giản. Các hiệu ứng chuyển động mượt mà hỗ trợ tốt cho cảm giác chơi (game feel)."
            elif "retention" in cr_lower or "giữ chân" in cr_lower:
                comment = f"Hệ thống nhiệm vụ hàng ngày, sự kiện tuần và cơ chế thưởng đăng nhập hoạt động hiệu quả để giữ chân người chơi lâu dài."
            elif "usp" in cr_lower or "điểm đặc biệt" in cr_lower:
                comment = f"Điểm độc đáo nằm ở việc tối ưu hóa hiệu năng cực tốt và kết hợp các yếu tố gameplay sáng tạo so với các đối thủ cùng thể loại trên thị trường."
            else:
                # Fallback for custom criteria
                comment = f"Tiêu chí {cr} được phân tích dựa trên thông tin mô tả game: '{args.info}'. Đánh giá chung cho thấy tiềm năng phát triển và phù hợp với xu hướng thể loại {args.genre}."
                
            # Determine score based on sentiment in info for this criterion
            base_score = 3.5
            comment_lower = comment.lower()
            # positive keywords
            if any(w in comment_lower or w in info_lower for w in ["tốt", "mượt", "đẹp", "xuất sắc", "great", "excellent", "good", "ấn tượng", "sáng tạo"]):
                base_score += 0.8
            # negative keywords
            if any(w in comment_lower or w in info_lower for w in ["tệ", "lỗi", "giật", "chậm", "bad", "issue", "poor", "lạm dụng", "khó chịu"]):
                base_score -= 0.8
            # custom adjustment
            score = round(min(5.0, max(1.0, base_score)), 1)
            framework[cr] = {"comment": comment, "score": score}
            
        agent.extracted_framework = framework
        agent.raw_content = " ".join([f"{k}: {v['comment']}" for k, v in framework.items()])
        
        steps = agent.research()
        scores = agent.generate_scorecard(framework)
        report_file = agent.write_slides_report(scores, steps, output_dir)
        
        print(f"SUCCESS:{report_file}")
        return

    print("=== AI Game Evaluation Agent ===")
    print(f"Input directory: {doc_dir}")
    print(f"Output directory: {output_dir}")
    
    # Scan for both .xlsx and .xls files
    xlsx_files = glob.glob(os.path.join(doc_dir, "*.xlsx"))
    xls_files = glob.glob(os.path.join(doc_dir, "*.xls"))
    
    all_files = xlsx_files + xls_files
    
    if not all_files:
        print("No Excel files found in the doc folder.")
        print("Please place your game evaluation Excel files in: workflow-nhan/doc/")
        return
        
    print(f"Found {len(all_files)} Excel file(s) to process.")
    
    processed_count = 0
    for file_path in all_files:
        filename = os.path.basename(file_path)
        # Skip temp files or metadata lock files
        if filename.startswith("~$") or filename.endswith(".temp.xlsx"):
            continue
            
        print(f"\nProcessing: {filename}")
        
        agent = GameEvalAgent()
        if agent.load_from_xlsx(file_path):
            framework = agent.extracted_framework
            steps = agent.research()
            scores = agent.generate_scorecard(framework)
            report_file = agent.write_slides_report(scores, steps, output_dir)
            print(f"Successfully processed {filename} -> {os.path.basename(report_file)}")
            processed_count += 1
        else:
            print(f"Failed to process: {filename}")
            
    print(f"\nProcessing completed. Successfully compiled {processed_count}/{len(all_files)} files.")

if __name__ == "__main__":
    main()
